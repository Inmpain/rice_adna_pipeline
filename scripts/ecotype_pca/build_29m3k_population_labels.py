#!/usr/bin/env python3
"""
Attach real 3K Rice Genomes Project population labels to NB_final_snp.ind.

The panel's .ind file currently carries a "???" placeholder in its
population column (confirmed on the server: `awk '{print $3}'
NB_final_snp.ind | sort -u` returns exactly one value, "???", for all
3,024 samples) -- there is no smartpca poplistname without this being
filled in first. This is docs/ECOTYPE_PCA_PANEL.md section 3.2 todo item 2.

SAMPLE-ID FAMILIES (confirmed against the real .ind file, not assumed):
  - "IRIS_313-XXXXX" (2,466 samples, exact 1:1 count match with Table S1A's
    row count) -> looked up in Table S1A ("Information for the 2,466 rice
    accessions from ... IRRI") by DNA_UNIQUE_ID, after normalizing the
    metadata's "IRIS 313-XXXXX" (space) to "IRIS_313-XXXXX" (underscore) --
    the underscore form is confirmed as the real sequencing-data convention
    via docs/references/3k_rice_genomes_project/seq_file_mapping_to_SRA.txt's
    Sample_alias column, not guessed.
  - "B###" and "CX#" (246 + ~312 samples respectively, both matched) ->
    both look up in Table S1B ("... 534 accessions from the China National
    Crop Genebank and CAAS") by DNA_UNIQUE_ID directly, no normalization
    needed -- Table S1B mixes multiple internal sub-sources under one
    sheet: "MC" rows use "B###" IDs, "IRMBN" rows use "CX#" IDs (289 such
    rows, confirmed by a full-column scan, not the first few rows). A
    same-panel sample ID overlap between NB_final_snp.ind's "CX#" IDs and
    asn720.6m.ind (the "CX382 duplicate suspicion", docs/ECOTYPE_PCA_PANEL.md
    section 3.2 todo item 4) is a separate, still-open question about
    whether the SAME physical accession was independently included in both
    panels -- unrelated to whether this script can label it, which it can.
  - Any sample_id with no metadata entry at all (of any prefix) gets
    --unmapped-label (default "UNK"); the coverage report distinguishes
    these from genuinely-matched-but-ambiguous labels.

LABEL STANDARDIZATION IS PARTIAL, DELIBERATELY:
The xlsx's "Variety Group (Tree)" column documents itself (Column-Info
sheet) only as "Variety group placement based on Neighbor Joining
analysis" -- this is the raw output of the *original 2014* NJ-tree
classification, which predates the 3K RGP paper's final published
IND/AUS/ARO/TRJ/TEJ/ADM nomenclature (Wang et al. 2018). Six of the eight
raw values found in the file map onto that scheme unambiguously (Indica,
Aus/Aus-boro, Basmati-sadri/Aromatic, Tropical japonica, Temperate
japonica). Two do NOT and are deliberately kept as their own labels
instead of being silently forced into the six-code scheme:
  - "Japonica" / "japonica" (unqualified, no tropical/temperate split)
    -> written as JAPONICA_UNSPEC.
  - "Intermediate type" -> written as INTERMEDIATE_TYPE. This is plausibly
    what later 3K releases call "Admixed", but this 2014 file never uses
    that word, so this script does not assert the equivalence. Confirm
    against the actual 2018 3K RGP paper's sample list before treating
    INTERMEDIATE_TYPE as ADM in a smartpca poplistname.

SAFETY: same discipline as merge_ancient_into_panel.py -- every check
below is a hard failure (non-zero exit, nothing written to --out) rather
than a warning: duplicate DNA_UNIQUE_ID within a metadata sheet, duplicate
sample ID within the .ind file, a .ind line that doesn't parse into
exactly 3 whitespace-separated fields, and an unrecognized (not one of
the eight known values) raw Variety Group string. Output is written to a
temp file in the same directory and atomically renamed into place
(os.replace) only after a complete, verified pass succeeds.

Usage:
  python3 build_29m3k_population_labels.py \\
    --ind /home/scratch/yinmt202607/db/29M_3k/NB_final_snp.ind \\
    --metadata-xlsx rice_line_metadata_20141029.xlsx \\
    --out /home/scratch/yinmt202607/db/29M_3k/NB_final_snp.labeled.ind \\
    --report /home/scratch/yinmt202607/db/29M_3k/NB_final_snp.label_report.tsv
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import tempfile
from collections import Counter
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover - exercised on cluster, not stdlib CI
    openpyxl = None


IRIS_ID_RE = re.compile(r"^IRIS_313-\d+$")
B_ID_RE = re.compile(r"^B\d+$")
CX_ID_RE = re.compile(r"^CX\d+$")

# Every raw "Variety Group (Tree)" value actually observed across both
# xlsx sheets (verified by a full-column scan, not assumed) maps to
# exactly one of these two dicts. Anything else is a hard error --
# see the module docstring for why the two AMBIGUOUS entries are not
# folded into the six-code CLEAN scheme.
CLEAN_LABELS: dict[str, str] = {
    "Indica": "IND",
    "Aus/boro": "AUS",
    "Aus": "AUS",
    "Basmati/sadri": "ARO",
    # This exact string (truncated mid-word) is what the xlsx cell contains.
    "Aromatic (basmati/sandri type": "ARO",
    "Tropical japonica": "TRJ",
    "Temperate japonica": "TEJ",
}
AMBIGUOUS_LABELS: dict[str, str] = {
    "Japonica": "JAPONICA_UNSPEC",
    "japonica": "JAPONICA_UNSPEC",
    "Intermediate type": "INTERMEDIATE_TYPE",
}

SHEET_SPECS = [
    # (sheet_name, header_row_1indexed, dna_unique_id_col0, variety_group_col0, normalize_id)
    ("Table S1A", 2, 2, 13, lambda s: s.replace(" ", "_")),
    ("Table S1B", 2, 2, 13, lambda s: s),
]


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--ind", required=True, help="path to NB_final_snp.ind")
    parser.add_argument("--metadata-xlsx", required=True, help="path to rice_line_metadata_20141029.xlsx")
    parser.add_argument("--out", required=True, help="path for the new .ind with real labels")
    parser.add_argument("--report", required=True, help="path for a per-sample coverage TSV")
    parser.add_argument(
        "--unmapped-label",
        default="UNK",
        help="population label for samples with no metadata match (default: UNK)",
    )
    return parser.parse_args(argv)


def validate_args(args: argparse.Namespace) -> None:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to read the xlsx metadata file")
    if not Path(args.ind).is_file():
        raise FileNotFoundError(f".ind file not found: {args.ind}")
    if not Path(args.metadata_xlsx).is_file():
        raise FileNotFoundError(f"metadata xlsx not found: {args.metadata_xlsx}")


def standardize_label(raw: str) -> str:
    if raw in CLEAN_LABELS:
        return CLEAN_LABELS[raw]
    if raw in AMBIGUOUS_LABELS:
        return AMBIGUOUS_LABELS[raw]
    raise ValueError(
        f"unrecognized Variety Group value {raw!r} -- not one of the "
        f"{len(CLEAN_LABELS) + len(AMBIGUOUS_LABELS)} values this script "
        "was built against. The xlsx may have been edited, or a new sheet "
        "added; update CLEAN_LABELS/AMBIGUOUS_LABELS after checking."
    )


def load_metadata(xlsx_path: Path) -> dict[str, tuple[str, str, str]]:
    """Return normalized_sample_id -> (sheet_name, raw_group, standardized_label)."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    lookup: dict[str, tuple[str, str, str]] = {}
    for sheet_name, header_row, id_col, group_col, normalize in SHEET_SPECS:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"expected sheet {sheet_name!r} not found in {xlsx_path}")
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            raw_id = row[id_col] if id_col < len(row) else None
            if raw_id is None:
                continue  # padding row past the sheet's real accession count
            raw_group = row[group_col] if group_col < len(row) else None
            sample_id = normalize(str(raw_id).strip())
            if sample_id in lookup:
                raise ValueError(
                    f"duplicate DNA_UNIQUE_ID {sample_id!r} in metadata "
                    f"(seen in {lookup[sample_id][0]!r} and {sheet_name!r})"
                )
            if raw_group is None:
                label = None
            else:
                label = standardize_label(str(raw_group).strip())
            lookup[sample_id] = (sheet_name, raw_group, label)
    return lookup


def classify_id(sample_id: str) -> str:
    if IRIS_ID_RE.match(sample_id):
        return "IRIS"
    if B_ID_RE.match(sample_id):
        return "B"
    if CX_ID_RE.match(sample_id):
        return "CX"
    return "OTHER"


def load_ind(ind_path: Path) -> list[tuple[str, str, str]]:
    rows: list[tuple[str, str, str]] = []
    seen: set[str] = set()
    with ind_path.open() as handle:
        for line_no, line in enumerate(handle, start=1):
            fields = line.split()
            if len(fields) != 3:
                raise ValueError(
                    f"{ind_path}:{line_no}: expected 3 whitespace-separated "
                    f"fields (sample sex label), got {len(fields)}: {line!r}"
                )
            sample_id, sex, old_label = fields
            if sample_id in seen:
                raise ValueError(f"duplicate sample ID in {ind_path}: {sample_id!r}")
            seen.add(sample_id)
            rows.append((sample_id, sex, old_label))
    return rows


def write_atomic(path: Path, write_fn) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=path.parent)
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        with tmp_path.open("w", newline="") as handle:
            write_fn(handle)
        os.replace(tmp_path, path)
    except BaseException:
        tmp_path.unlink(missing_ok=True)
        raise


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        validate_args(args)
        metadata = load_metadata(Path(args.metadata_xlsx))
        ind_rows = load_ind(Path(args.ind))

        report_rows = []
        new_ind_lines = []
        label_counts: Counter[str] = Counter()
        class_matched: Counter[str] = Counter()
        class_total: Counter[str] = Counter()

        for sample_id, sex, _old_label in ind_rows:
            id_class = classify_id(sample_id)
            class_total[id_class] += 1
            meta = metadata.get(sample_id)
            if meta is None or meta[2] is None:
                final_label = args.unmapped_label
                source_sheet = meta[0] if meta is not None else ""
                raw_group = meta[1] if meta is not None else ""
            else:
                source_sheet, raw_group, final_label = meta
                class_matched[id_class] += 1
            label_counts[final_label] += 1
            report_rows.append(
                {
                    "sample_id": sample_id,
                    "id_class": id_class,
                    "source_sheet": source_sheet,
                    "raw_variety_group": raw_group if raw_group is not None else "",
                    "final_label": final_label,
                }
            )
            new_ind_lines.append(f"{sample_id:>18} {sex} {final_label}\n")

        write_atomic(Path(args.out), lambda h: h.writelines(new_ind_lines))

        def write_report(handle):
            writer = csv.DictWriter(
                handle,
                delimiter="\t",
                fieldnames=["sample_id", "id_class", "source_sheet", "raw_variety_group", "final_label"],
            )
            writer.writeheader()
            writer.writerows(report_rows)

        write_atomic(Path(args.report), write_report)

    except (OSError, ValueError, RuntimeError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"[labels] total samples: {len(ind_rows)}", file=sys.stderr)
    for id_class in ("IRIS", "B", "CX", "OTHER"):
        total = class_total.get(id_class, 0)
        if total == 0:
            continue
        matched = class_matched.get(id_class, 0)
        print(f"[labels]   {id_class}: {matched}/{total} matched a metadata row", file=sys.stderr)
    print("[labels] final label distribution:", file=sys.stderr)
    for label, count in label_counts.most_common():
        print(f"[labels]   {label}: {count}", file=sys.stderr)
    print(f"[done] wrote {args.out}", file=sys.stderr)
    print(f"[done] wrote {args.report}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
